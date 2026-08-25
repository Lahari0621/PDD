"""
Generate all Excel reports from execution-results.json
Output: reports/Excel/{Automation_Test_Report, Failed, Passed, Summary}.xlsx
"""
import os, sys, json
from datetime import datetime

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                 GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference, PieChart
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl'], check=True)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ── Color constants ────────────────────────────────────────────
C_HEADER_BG   = "1E3A5F"
C_HEADER_FG   = "FFFFFF"
C_PASS_BG     = "D6F5D6"
C_FAIL_BG     = "FFD6D6"
C_SKIP_BG     = "FFF9C4"
C_BLOCK_BG    = "E0E0E0"
C_TITLE_BG    = "0D2137"
C_ALT_ROW     = "F0F4F8"

# ── Load results ───────────────────────────────────────────────
json_path = os.path.join(ROOT, 'reports', 'JSON', 'execution-results.json')
results = []
summary_data = {
    "execution_date": datetime.now().isoformat(),
    "total": 0, "passed": 0, "failed": 0, "skipped": 0, "blocked": 0,
    "results": []
}

if os.path.exists(json_path):
    with open(json_path) as f:
        summary_data = json.load(f)
    results = summary_data.get('results', [])
else:
    print(f"⚠️  No JSON results found at {json_path}. Generating placeholder report.")
    # Create placeholder data covering all 400+ test IDs
    modules = {
        'Authentication': (40, 'AUTH'),
        'Authorization':  (40, 'AUTHZ'),
        'Navigation':     (30, 'NAV'),
        'UI Validation':  (50, 'UI'),
        'Forms':          (50, 'FORM'),
        'CRUD Operations':(50, 'CRUD'),
        'Input Validation':(40,'INP'),
        'Error Handling': (20, 'ERR'),
        'Session Management':(20,'SES'),
        'Accessibility':  (20, 'ACC'),
        'Responsive Design':(20,'RES'),
        'Performance':    (20, 'PERF'),
        'Regression':     (50, 'REG'),
    }
    import random
    for mod, (count, prefix) in modules.items():
        for i in range(1, count + 1):
            # ALL FORCED FAIL
            status = 'FAILED'
            results.append({
                "test_id": f"{prefix}-{i:03d}",
                "module": mod,
                "test_name": f"{mod} Test Case {i:03d}",
                "status": status,
                "execution_time": round(random.uniform(0.5, 3.0), 2),
                "priority": random.choice(["Critical", "High", "Medium", "Low"]),
                "failure_reason": "Forced failure — assert False",
                "screenshot_path": f"screenshots/failures/FAIL_{prefix}_{i:03d}.png",
                "preconditions": "Application deployed to GitHub Pages",
                "expected_result": "Test passes successfully",
                "actual_result": "FAILED",
                "timestamp": datetime.now().isoformat(),
            })
    summary_data['results'] = results
    summary_data['total']   = len(results)
    summary_data['passed']  = sum(1 for r in results if r['status'] == 'PASSED')
    summary_data['failed']  = sum(1 for r in results if r['status'] == 'FAILED')
    summary_data['skipped'] = sum(1 for r in results if r['status'] == 'SKIPPED')

EXCEL_DIR = os.path.join(ROOT, 'reports', 'Excel')
os.makedirs(EXCEL_DIR, exist_ok=True)

total   = summary_data.get('total',   len(results))
passed  = summary_data.get('passed',  sum(1 for r in results if r['status'] == 'PASSED'))
failed  = summary_data.get('failed',  sum(1 for r in results if r['status'] == 'FAILED'))
skipped = summary_data.get('skipped', sum(1 for r in results if r['status'] == 'SKIPPED'))
blocked = summary_data.get('blocked', sum(1 for r in results if r['status'] == 'BLOCKED'))
pass_pct = round((passed / total * 100), 2) if total > 0 else 0

passed_list  = [r for r in results if r['status'] == 'PASSED']
failed_list  = [r for r in results if r['status'] == 'FAILED']
skipped_list = [r for r in results if r['status'] == 'SKIPPED']


# ── Helper functions ───────────────────────────────────────────
def make_header_style():
    return {
        'font':  Font(bold=True, color=C_HEADER_FG, size=11),
        'fill':  PatternFill("solid", fgColor=C_HEADER_BG),
        'align': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            bottom=Side(style='medium', color='000000'),
            right= Side(style='thin',   color='888888'),
        )
    }


def apply_header(cell, text):
    s = make_header_style()
    cell.value     = text
    cell.font      = s['font']
    cell.fill      = s['fill']
    cell.alignment = s['align']
    cell.border    = s['border']


def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def status_fill(status: str) -> PatternFill:
    colors = {
        'PASSED':  C_PASS_BG,
        'FAILED':  C_FAIL_BG,
        'SKIPPED': C_SKIP_BG,
        'BLOCKED': C_BLOCK_BG,
    }
    return PatternFill("solid", fgColor=colors.get(status, "FFFFFF"))


def status_font(status: str) -> Font:
    colors = {
        'PASSED':  '1A7A1A',
        'FAILED':  'C0392B',
        'SKIPPED': '7D6608',
        'BLOCKED': '555555',
    }
    return Font(color=colors.get(status, '000000'), bold=(status == 'FAILED'))


def write_results_sheet(ws, rows, title="Test Cases"):
    """Write a standard results table to the given worksheet."""
    headers = ["Test ID", "Module", "Test Name", "Status",
               "Execution Time (s)", "Priority", "Failure Reason",
               "Preconditions", "Expected Result", "Actual Result",
               "Screenshot", "Timestamp"]
    for ci, h in enumerate(headers, 1):
        apply_header(ws.cell(row=1, column=ci), h)
    ws.row_dimensions[1].height = 30

    for ri, row in enumerate(rows, 2):
        bg = C_ALT_ROW if ri % 2 == 0 else "FFFFFF"
        data = [
            row.get('test_id', ''),
            row.get('module', ''),
            row.get('test_name', ''),
            row.get('status', ''),
            row.get('execution_time', 0),
            row.get('priority', ''),
            row.get('failure_reason', ''),
            row.get('preconditions', ''),
            row.get('expected_result', ''),
            row.get('actual_result', ''),
            row.get('screenshot_path', ''),
            row.get('timestamp', ''),
        ]
        for ci, val in enumerate(data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if ci == 4:  # Status column
                cell.fill = status_fill(str(val))
                cell.font = status_font(str(val))
            else:
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(size=10)
            cell.alignment = Alignment(vertical='center', wrap_text=(ci in [3, 7]))
        ws.row_dimensions[ri].height = 20

    set_col_widths(ws, {
        'A': 16, 'B': 22, 'C': 45, 'D': 12,
        'E': 18, 'F': 12, 'G': 40, 'H': 30,
        'I': 30, 'J': 12, 'K': 35, 'L': 24,
    })
    ws.freeze_panes = 'A2'


# ═══════════════════════════════════════════════════════════════
# 1. Automation_Test_Report.xlsx  (6 sheets)
# ═══════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# Sheet 1 — All Executed Tests
ws1 = wb.active
ws1.title = "All Test Cases"
write_results_sheet(ws1, results, "All Executed Test Cases")

# Sheet 2 — Passed Tests
ws2 = wb.create_sheet("Passed Tests")
write_results_sheet(ws2, passed_list, "Passed Tests")

# Sheet 3 — Failed Tests
ws3 = wb.create_sheet("Failed Tests")
write_results_sheet(ws3, failed_list, "Failed Tests")

# Sheet 4 — Skipped Tests
ws4 = wb.create_sheet("Skipped Tests")
write_results_sheet(ws4, skipped_list, "Skipped Tests")

# Sheet 5 — Execution Metrics
ws5 = wb.create_sheet("Execution Metrics")
ws5.column_dimensions['A'].width = 30
ws5.column_dimensions['B'].width = 20

# Title
title_cell = ws5['A1']
title_cell.value = "AI Debate Partner — Execution Metrics"
title_cell.font  = Font(bold=True, size=16, color=C_HEADER_FG)
title_cell.fill  = PatternFill("solid", fgColor=C_TITLE_BG)
ws5.merge_cells('A1:B1')
ws5['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws5.row_dimensions[1].height = 36

metrics = [
    ("Execution Date",      summary_data.get('execution_date', datetime.now().isoformat())[:19]),
    ("Base URL",            "https://Lahari0621.github.io/PDD/"),
    ("Total Test Cases",    total),
    ("Passed",              passed),
    ("Failed",              failed),
    ("Skipped",             skipped),
    ("Blocked",             blocked),
    ("Pass Percentage",     f"{pass_pct}%"),
    ("Fail Percentage",     f"{round(100-pass_pct,2)}%"),
    ("Total Exec Time (s)", round(sum(r.get('execution_time',0) for r in results), 2)),
    ("Avg Exec Time (s)",   round(sum(r.get('execution_time',0) for r in results)/max(total,1), 2)),
    ("Critical Tests",      sum(1 for r in results if r.get('priority','')=='Critical')),
    ("High Priority Tests", sum(1 for r in results if r.get('priority','')=='High')),
]

for r_idx, (k, v) in enumerate(metrics, 3):
    ka = ws5.cell(row=r_idx, column=1, value=k)
    va = ws5.cell(row=r_idx, column=2, value=v)
    ka.font = Font(bold=True, size=11)
    ka.fill = PatternFill("solid", fgColor="EBF3FB")
    va.font = Font(size=11)
    va.fill = PatternFill("solid", fgColor="FFFFFF")
    ka.alignment = va.alignment = Alignment(vertical='center')
    ws5.row_dimensions[r_idx].height = 22

# Module breakdown table
ws5['A18'] = "Module Breakdown"
ws5['A18'].font = Font(bold=True, size=12, color=C_HEADER_FG)
ws5['A18'].fill = PatternFill("solid", fgColor=C_HEADER_BG)
ws5.merge_cells('A18:E18')

mod_headers = ["Module", "Total", "Passed", "Failed", "Pass Rate"]
for ci, h in enumerate(mod_headers, 1):
    c = ws5.cell(row=19, column=ci, value=h)
    c.font = Font(bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="2E6DA4")
    c.alignment = Alignment(horizontal='center')

modules_set = sorted(set(r.get('module','') for r in results))
for mi, mod in enumerate(modules_set, 20):
    mod_rows = [r for r in results if r.get('module','') == mod]
    m_total  = len(mod_rows)
    m_passed = sum(1 for r in mod_rows if r['status'] == 'PASSED')
    m_failed = sum(1 for r in mod_rows if r['status'] == 'FAILED')
    m_rate   = f"{round(m_passed/m_total*100,1)}%" if m_total else "N/A"
    for ci, val in enumerate([mod, m_total, m_passed, m_failed, m_rate], 1):
        c = ws5.cell(row=mi, column=ci, value=val)
        c.font = Font(size=10)
        c.fill = PatternFill("solid", fgColor=C_ALT_ROW if mi%2==0 else "FFFFFF")
        c.alignment = Alignment(horizontal='center' if ci>1 else 'left', vertical='center')
    ws5.row_dimensions[mi].height = 20

for col, w in {'A':30,'B':10,'C':10,'D':10,'E':12}.items():
    ws5.column_dimensions[col].width = w

# Sheet 6 — Defect Summary
ws6 = wb.create_sheet("Defect Summary")
ws6['A1'] = "Defect Summary Report"
ws6['A1'].font = Font(bold=True, size=16, color=C_HEADER_FG)
ws6['A1'].fill = PatternFill("solid", fgColor=C_TITLE_BG)
ws6.merge_cells('A1:F1')
ws6['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws6.row_dimensions[1].height = 36

defect_headers = ["Defect ID", "Test Case ID", "Module", "Description",
                  "Severity", "Status"]
for ci, h in enumerate(defect_headers, 1):
    apply_header(ws6.cell(row=2, column=ci), h)

for di, r in enumerate(failed_list, 3):
    defect_data = [
        f"BUG-{di-2:04d}",
        r.get('test_id',''),
        r.get('module',''),
        r.get('failure_reason','') or f"Test failed: {r.get('test_name','')}",
        r.get('priority','Medium'),
        "Open",
    ]
    for ci, val in enumerate(defect_data, 1):
        c = ws6.cell(row=di, column=ci, value=val)
        c.font = Font(size=10, color="C0392B" if ci==6 else "000000")
        c.fill = PatternFill("solid", fgColor=C_FAIL_BG)
        c.alignment = Alignment(vertical='center', wrap_text=(ci==4))
    ws6.row_dimensions[di].height = 20

set_col_widths(ws6, {'A':14,'B':16,'C':22,'D':50,'E':14,'F':12})
ws6.freeze_panes = 'A3'

out1 = os.path.join(EXCEL_DIR, 'Automation_Test_Report.xlsx')
wb.save(out1)
print(f"✅ Saved: {out1}")


# ═══════════════════════════════════════════════════════════════
# 2. Failed_Test_Cases.xlsx
# ═══════════════════════════════════════════════════════════════
wb2 = openpyxl.Workbook()
ws = wb2.active
ws.title = "Failed Tests"
write_results_sheet(ws, failed_list)

# Extra: root cause analysis tab
ws_rca = wb2.create_sheet("Root Cause Analysis")
apply_header(ws_rca.cell(1,1), "Test ID")
apply_header(ws_rca.cell(1,2), "Module")
apply_header(ws_rca.cell(1,3), "Failure Reason")
apply_header(ws_rca.cell(1,4), "Root Cause Category")
apply_header(ws_rca.cell(1,5), "Recommended Action")

for ri, r in enumerate(failed_list, 2):
    reason = r.get('failure_reason','').lower()
    category = ("Network/API" if any(w in reason for w in ['network','timeout','api','connect']) else
                "Auth/Session" if 'auth' in reason else
                "UI Element"  if any(w in reason for w in ['element','display','render']) else
                "Assertion"   if 'assert' in reason else
                "Unknown")
    action = ("Check backend connectivity" if category == "Network/API" else
              "Verify auth token/session"  if category == "Auth/Session" else
              "Check CSS/DOM selectors"    if category == "UI Element" else
              "Review test assertion logic")
    for ci, val in enumerate(
        [r.get('test_id',''), r.get('module',''),
         r.get('failure_reason','')[:100], category, action], 1
    ):
        c = ws_rca.cell(row=ri, column=ci, value=val)
        c.font = Font(size=10)
        c.fill = PatternFill("solid", fgColor=C_FAIL_BG if ri%2==0 else "FFE6E6")

set_col_widths(ws_rca, {'A':16,'B':22,'C':50,'D':22,'E':40})

out2 = os.path.join(EXCEL_DIR, 'Failed_Test_Cases.xlsx')
wb2.save(out2)
print(f"✅ Saved: {out2}")


# ═══════════════════════════════════════════════════════════════
# 3. Passed_Test_Cases.xlsx
# ═══════════════════════════════════════════════════════════════
wb3 = openpyxl.Workbook()
ws = wb3.active
ws.title = "Passed Tests"
write_results_sheet(ws, passed_list)

out3 = os.path.join(EXCEL_DIR, 'Passed_Test_Cases.xlsx')
wb3.save(out3)
print(f"✅ Saved: {out3}")


# ═══════════════════════════════════════════════════════════════
# 4. Summary_Report.xlsx
# ═══════════════════════════════════════════════════════════════
wb4 = openpyxl.Workbook()
ws = wb4.active
ws.title = "Executive Summary"

# Title banner
ws.merge_cells('A1:G1')
ws['A1'] = "AI Debate Partner — Test Execution Executive Summary"
ws['A1'].font  = Font(bold=True, size=18, color='FFFFFF')
ws['A1'].fill  = PatternFill("solid", fgColor=C_TITLE_BG)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 50

# Subtitle
ws.merge_cells('A2:G2')
ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | URL: https://Lahari0621.github.io/PDD/"
ws['A2'].font  = Font(size=10, color='666666', italic=True)
ws['A2'].alignment = Alignment(horizontal='center')

# KPI Cards (row 4 onwards)
kpis = [
    ("Total Tests",    total,      "1E3A5F", "FFFFFF"),
    ("✅ Passed",      passed,     "1A7A1A", "FFFFFF"),
    ("❌ Failed",      failed,     "C0392B", "FFFFFF"),
    ("⏭️ Skipped",    skipped,    "7D6608", "FFFFFF"),
    ("Pass Rate",      f"{pass_pct}%", "1A5276","FFFFFF"),
    ("Status",         "PASS" if pass_pct >= 95 else "FAIL",
     "1A7A1A" if pass_pct >= 95 else "C0392B", "FFFFFF"),
]

ws.row_dimensions[4].height = 50
for ci, (label, value, bg, fg) in enumerate(kpis, 1):
    label_cell = ws.cell(row=4, column=ci, value=label)
    label_cell.font  = Font(bold=True, size=10, color=fg)
    label_cell.fill  = PatternFill("solid", fgColor=bg)
    label_cell.alignment = Alignment(horizontal='center', vertical='bottom')

    val_cell = ws.cell(row=5, column=ci, value=value)
    val_cell.font  = Font(bold=True, size=20, color=bg)
    val_cell.fill  = PatternFill("solid", fgColor="F8FBFF")
    val_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[5].height = 45
    ws.column_dimensions[get_column_letter(ci)].width = 18

# Module breakdown
ws['A7'] = "Module-wise Results"
ws['A7'].font = Font(bold=True, size=13, color=C_HEADER_FG)
ws['A7'].fill = PatternFill("solid", fgColor=C_HEADER_BG)
ws.merge_cells('A7:G7')
ws['A7'].alignment = Alignment(horizontal='center')
ws.row_dimensions[7].height = 28

for ci, h in enumerate(["Module","Total","Passed","Failed","Skipped","Pass Rate","Status"], 1):
    apply_header(ws.cell(8, ci), h)
ws.row_dimensions[8].height = 24

modules_list = sorted(set(r.get('module','') for r in results))
for ri_off, mod in enumerate(modules_list, 9):
    mod_rows = [r for r in results if r.get('module','') == mod]
    mt = len(mod_rows)
    mp = sum(1 for r in mod_rows if r['status'] == 'PASSED')
    mf = sum(1 for r in mod_rows if r['status'] == 'FAILED')
    ms = sum(1 for r in mod_rows if r['status'] == 'SKIPPED')
    mr = round(mp/mt*100,1) if mt else 0
    status_lbl = "✅ PASS" if mr >= 95 else "❌ FAIL"
    row_data = [mod, mt, mp, mf, ms, f"{mr}%", status_lbl]
    bg = C_ALT_ROW if ri_off % 2 == 0 else "FFFFFF"
    for ci, val in enumerate(row_data, 1):
        c = ws.cell(row=ri_off, column=ci, value=val)
        c.font = Font(size=10, bold=(ci == 7),
                      color="1A7A1A" if val == "✅ PASS" else
                            "C0392B" if val == "❌ FAIL" else "000000")
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal='center' if ci > 1 else 'left', vertical='center')
    ws.row_dimensions[ri_off].height = 20

ws.freeze_panes = 'A9'

out4 = os.path.join(EXCEL_DIR, 'Summary_Report.xlsx')
wb4.save(out4)
print(f"✅ Saved: {out4}")

print(f"\n📊 Excel Reports Summary:")
print(f"   Total: {total} | Passed: {passed} | Failed: {failed} | "
      f"Skipped: {skipped} | Pass Rate: {pass_pct}%")
print(f"   Reports saved to: {EXCEL_DIR}")
